from openpyxl import load_workbook


def get_row_count(path, sheet_name):
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(path, sheet_name):
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def get_cell_data(path, sheet_name, row_number, column_number):
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value


def set_cell_data(path, sheet_name, row_number, column_number, data):
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(path)


def get_data_from_excel(path, sheet_name):
    final_list = []
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_column = sheet.max_column

    for i in range(2, total_rows + 1):
        row_list = []
        for j in range(1, total_column + 1):
            row_list.append(sheet.cell(row=i, column=j).value)
        final_list.append(row_list)

    return final_list
